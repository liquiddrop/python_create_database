# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a script file to parse wikipedia for historical events.
"""
import wikipedia
import urllib.request
from bs4 import BeautifulSoup
import pandas
from openpyxl import load_workbook
import random
import re
import requests
import json
import os

#FILTER_LENGTH is the max amount of characters allowed before ':' to still
#count as a historical event this is based on the format "date:event" 
FILTER_LENGTH=10
#SUMMARY_LENGTH is the number of sentences to take from the wikipedia summary
#to output to the database file
SUMMARY_LENGTH=4
#output xml and pic names
OUTPUT_FILE=r'C:\Users\zerol\Documents\context_database\historical_event_database.xlsx'
OUTPUT_DIR=r'C:\Users\zerol\Documents\context_database\\'
DEBUG_PRINT=False
DEBUG_PRINT_FULL=False
DEBUG_TIMING=False
WIKI_REQUEST='http://en.wikipedia.org/w/api.php?action=query&prop=pageimages&format=json&piprop=original&titles='

def get_wiki_image(title):
    try:
        response=requests.get(WIKI_REQUEST+title)
        json_data=json.loads(response.text)
        img_link=list(json_data['query']['pages'].values())[0]['original']['source']
        return img_link        
    except:
        return 0

#return a wikipedia page object based on the title if we can't get it just
#take a random page from the suggestions
def get_wiki_page(title):
    try:
        page_image=wikipedia.page(title, auto_suggest=False)
    except wikipedia.DisambiguationError as e:
        try:
            s = random.choice(e.options)
            if DEBUG_PRINT:
                print("Handling disambiguation error and options are")
                print(e.options)
                print(s)
            #return None
            page_image=wikipedia.page(s, auto_suggest=False)
        except:
            #if this returns two disambiguation errors then it is too general
            #so stop trying and return none
            print("We are in the second disambiguation error so return none for:")
            print(s)
            return None

    return page_image

#given a wikipedia page object get the first image if it exists
def get_wiki_picture(page_image,pic_name):
    image_down_link=get_wiki_image(page_image.title)
    if DEBUG_PRINT:
        print(page_image.title)
        print(image_down_link)
    if image_down_link:
        image_type=os.path.splitext(image_down_link)
        urllib.request.urlretrieve(image_down_link , OUTPUT_DIR + pic_name + image_type[1])
        return image_type[1]
    return ".blank"

#given a search able string search for wiki data and return in a list
#[pic_name, link, full_summary]
def get_wiki_data(temp_name):
    output_list=[]
    #need to limit the search string based off the limit of wikipedia
    #to only take a search under 300 make it 299 just to be safe 
    if(len(temp_name)>299):
        temp_name=temp_name[:299]
    
    temp_time=time()
    suggested_list=wikipedia.search(temp_name)
    if DEBUG_TIMING:
        print(f'wiki search time is {time() - temp_time} seconds')
    if DEBUG_PRINT:
        print(suggested_list)
        print(len(suggested_list))
    
    #check to returned suggested list to make sure it has content
    if (len(suggested_list) > 0):
        first_suggestion=suggested_list[0]
        #try to not take a timeline page but instead the full content page
        if ('timeline' in first_suggestion.lower()):
            #if timeline is in the first suggestion just try the next one
            if len(suggested_list) > 2:
                pic_name=suggested_list[1].replace(" ", "_")
                wiki_page=get_wiki_page(suggested_list[1])
                if wiki_page == None:
                    #need to return none so that the lengths match up
                    return [None, None, None]
                temp_time=time()
                ext_str=get_wiki_picture(wiki_page, pic_name)
                if DEBUG_TIMING:
                    print(f'get wiki pic time is {time() - temp_time} seconds')
                pic_name=pic_name + ext_str
                if DEBUG_PRINT:
                    print(pic_name)
                    print(wiki_page.url)
                output_list.append(pic_name)
                output_list.append(wiki_page.url)
                page_summary=wiki_page.summary
                short_summary=' '.join(re.split(r'(?<=[.:;])\s', page_summary)[:SUMMARY_LENGTH])                      
                output_list.append(short_summary)
                if DEBUG_PRINT_FULL:
                    print(page_summary)
                    print(short_summary)
            else:
                #need to return none so that the lengths match up
                return [None, None, None]
        else:
            pic_name=suggested_list[0].replace(" ", "_")
            wiki_page=get_wiki_page(suggested_list[0])
            if wiki_page == None:
                #need to return none so that the lengths match up
                return [None, None, None]
            temp_time=time()
            ext_str=get_wiki_picture(wiki_page, pic_name)
            if DEBUG_TIMING:
                print(f'wiki get pic time is {time() - temp_time} seconds')
            pic_name=pic_name + ext_str
            if DEBUG_PRINT:
                print(pic_name)
                print(wiki_page.url)
            output_list.append(pic_name)
            output_list.append(wiki_page.url)
            page_summary=wiki_page.summary
            short_summary=' '.join(re.split(r'(?<=[.:;])\s', page_summary)[:SUMMARY_LENGTH])
            output_list.append(short_summary)  
            if DEBUG_PRINT_FULL:
                print(page_summary)
                print(short_summary)
    else:
        #need to return none if the suggested list was empty
        return [None, None, None]
        
    return output_list

def input_to_dataframe(date=None, year=None, event_name=None, full_summary=None, picture=None, link=None):
    import pandas as pd
    df = pd.DataFrame({'date_string': [] if date==None else date})
    df['year']= year
    df['month']= ''
    df['day']= ''
    df['location_country']= ''
    df['location_state']= ''
    df['location_city']= ''
    df['importance']= ''
    df['summary']= event_name
    df['full_summary']= full_summary
    df['picture']= picture
    df['link']= link    
    return df

#output a dataframe to a excel file, if it doesn't exist create it
def output_dataframe_to_file(df):
    from os import path
    if not path.exists(OUTPUT_FILE):
        import pandas as pd
        # dataframe init based on the required database format
        blank_df=pd.DataFrame({'date_string': [],
                   'year' : [],
                   'month' : [],
                   'day' : [],
                   'location_country' : [],
                   'location_state' : [],
                   'location_city' : [],
                   'importance' : [],
                   'summary' : [],
                   'full_summary' : [],
                   'picture' : [],
                   'link' : []})
        writer=pd.ExcelWriter(OUTPUT_FILE, engine='xlsxwriter')
        blank_df.to_excel(writer,sheet_name="sheet1", startrow=0, index = False,header= True)
        writer.save()

    #append the datafame if the file already exists
    writer=pandas.ExcelWriter(OUTPUT_FILE, engine='openpyxl')
    book=load_workbook(OUTPUT_FILE)
    writer.book=book
    writer.sheets={ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

    writer.save()

#given a url parse historical events based of a table format
def parse_history_from_table_url(url):
    start=time()
    # open the url using urllib.request and put the HTML into the page variable
    page=urllib.request.urlopen(url)  

    # parse the HTML from our URL into the BeautifulSoup parse tree format
    soup=BeautifulSoup(page, "lxml")

    events_tables=soup.find_all("table", class_="wikitable")
    
    year=[]
    date=[]
    event_name=[]
    significance=[]
    picture=[]
    link=[]
    
    for table in events_tables:
        for row in table.findAll('tr'):
            cells=row.findAll('td')
            if len(cells)==4:
                year.append(cells[0].getText())
                date.append(cells[1].getText())
                event_name.append(cells[2].getText())
                significance.append(cells[3].getText())
                temp_time=time()
                wiki_data=get_wiki_data(cells[2].getText())
                if DEBUG_TIMING:
                    print(f'get_wiki_data took {time()-temp_time}')
                picture.append(wiki_data[0])
                link.append(wiki_data[1])

    temp_time=time()
    #create dataframe and input saved data
    df=input_to_dataframe(date, year,event_name,significance,picture,link)
    if DEBUG_TIMING:
        print(f'input to data frame took {time()-temp_time} seconds')
    temp_time=time()
    output_dataframe_to_file(df)

    if DEBUG_TIMING:
        print(f'output to dataframe took {time() - temp_time} seconds')
        print(f'Full time of function parse_from_table_url is {time() - start} seconds')

from time import time
#given a url parse historical events based off list format
def parse_history_from_list_url(url):
    start=time()
    
    # open the url using urllib.request and put the HTML into the page variable
    page=urllib.request.urlopen(url)

    # parse the HTML from our URL into the BeautifulSoup parse tree format
    soup=BeautifulSoup(page, "lxml")

    year=[]
    event_name=[]
    picture=[]
    link=[]
    full_summary=[]


    for tag in soup.findAll("li", attrs={'class': None}):
        if (':' in tag.get_text()):
            if(tag.text.find(':') <= FILTER_LENGTH):
                splitList=tag.text.split(':', 1)
                if DEBUG_PRINT:
                    print(splitList[0])
                    print(splitList[1][1:])
                year.append(splitList[0])
                temp_name=splitList[1][1:]
                event_name.append(temp_name)
                temp_time=time()
                wiki_data=get_wiki_data(temp_name)
                #print(f'wiki get data time {time() - temp_time} seconds')
                picture.append(wiki_data[0])
                link.append(wiki_data[1])
                full_summary.append(wiki_data[2])
                            
    temp_time=time()
    #create the dataframe and output to file
    df=input_to_dataframe(None, year,event_name,full_summary,picture,link)
    if DEBUG_TIMING:
        print(f'input to dataframe {time() - temp_time} seconds')

    temp_time=time()
    output_dataframe_to_file(df)
    if DEBUG_TIMING:
        print(f'output dataframe to file {time()-temp_time} seconds')
        end=time()
        print(print(f'The function took {end - start} seconds!'))

#parse historical events from list where the header is the year
def parse_history_from_list_with_header_url(url):
    start=time()
    # open the url using urllib.request and put the HTML into the page variable
    page=urllib.request.urlopen(url)

    # parse the HTML from our URL into the BeautifulSoup parse tree format
    soup=BeautifulSoup(page, "lxml")

    year=[]
    event_name=[]
    picture=[]
    link=[]
    full_summary=[]

    for header in soup.findAll("h3", attrs={'id': None}):
        temp_year=re.findall(r'\d+', header.text)
        for sib in header.find_next_siblings():
            if (sib.name == "ul"):
                event_list=sib.text
                for event in event_list.splitlines():
                    if DEBUG_PRINT:
                        print(temp_year[0])
                        print(event)
                    year.append(temp_year[0])
                    temp_name=event
                    event_name.append(temp_name)
                    temp_time=time()
                    wiki_data=get_wiki_data(temp_name)
                    if DEBUG_TIMING:
                        print(f'get_wiki_data took {time()-temp_time} seconds')
                    picture.append(wiki_data[0])
                    link.append(wiki_data[1])
                    full_summary.append(wiki_data[2])
            else:
                #not a ul so break and search for the next header
                break                            
    temp_time=time()
    #create the dataframe and output to file
    df=input_to_dataframe(None, year,event_name,full_summary,picture,link)
    if DEBUG_TIMING:
        print(f'input to dataframe took {time()-temp_time} seconds')
        temp_time=time()
    output_dataframe_to_file(df)
    if DEBUG_TIMING:
        print(f'output to dataframe took {time()-temp_time} seconds')
        print(f'total parse from list with header time is {time()-start} seconds')

#start the main function/call parsing on all the timelines
    
#be nice and don't request to much
wikipedia.set_rate_limiting(True)
wikipedia.set_lang('en')

print("Starting to parse the websites")
#Put list of wikiedia timeline pages followed by the parse here:
url = "https://en.wikipedia.org/wiki/Timeline_of_ancient_history"

parse_history_from_list_url(url)

url = "https://en.wikipedia.org/wiki/Timeline_of_the_Middle_Ages"

parse_history_from_table_url(url)

url = r"https://en.wikipedia.org/wiki/16th_century#Events"

parse_history_from_list_url(url)

url = r"https://en.wikipedia.org/wiki/Timeline_of_the_17th_century"

parse_history_from_list_url(url)

url = r"https://en.wikipedia.org/wiki/Timeline_of_the_18th_century"

parse_history_from_list_url(url)

url = r"https://en.wikipedia.org/wiki/Timeline_of_the_19th_century"

parse_history_from_list_url(url)

url = r"https://en.wikipedia.org/wiki/Timeline_of_the_20th_century"

parse_history_from_list_with_header_url(url)

url = r"https://en.wikipedia.org/wiki/Timeline_of_the_20th_century#1945"

parse_history_from_list_with_header_url(url)

url = r"https://en.wikipedia.org/wiki/Timeline_of_the_21st_century"

parse_history_from_list_with_header_url(url)

print("Finished parsing the last website")